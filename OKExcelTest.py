# 2025-01-07更新
# 从Excel文件读取URL，并解析出线路、单仓和多仓，并输出到Excel文件
import openpyxl
import requests
import json
import os
import re
import warnings 
warnings.filterwarnings("ignore")

def respon_out(respon): #输入respond,判断线路、单仓和多仓
    mcang=re.search("storeHouse",respon,re.M|re.I)
    scang = re.search("urls",respon,re.M|re.I)
    xlu = re.search("sites",respon,re.M|re.I)
    if xlu is not None :
        return "H线路"
    elif xlu is  None and scang is not None and mcang is None:
        return "H单仓"
    elif xlu is  None and scang is  None and mcang is not None:
        return "H多仓"
    else :
        return "other"

def addlink(url): #增加代理链接
    agentlink="https://xn--sss604efuw.com/jm/jiemi.php?url="
    url1=agentlink+url
    return url1

def url_respon(url): #输入原始url,解析网站。返回：0-代理url；1-respone等元组；2-状态+200
    headers={"User-Agent":"okhttp/4.1.0"}
    try:
        respon=requests.get(url,verify=False,timeout =60)
        codema= respon.status_code
        if codema==200:
            if respon_out(respon.text)!='other': #如果解析结果不是other，则返回解析结果
                conten="首次解析状态:"+str(codema)
                outcang=respon_out(respon.text)
                # print(respon.text)
                return url,conten,outcang
            else:
                respon1=requests.get(url,headers=headers,verify=False,timeout =60)  #更换UA
                codema1= respon1.status_code
                conten1="更换UA状态: "+str(codema1)
                outcang1=respon_out(respon1.text)
                return url,conten1,outcang1
        else:
            respon2=requests.get(url=addlink(url),headers=headers,verify=False,timeout =60)  #增加代理2函数
            codema2= respon.status_code
            conten2="解析状态: "+str(codema2)
            outcang2=respon_out(respon2.text)
            return url,conten2,outcang2
    except Exception:
        # print("解析失败了!: ",url)
        return url,"失败状态：000","解析失败: "
        
def main():
    i=1
    # excel_file = r'D:\OneDrive\Learn\pythonLearn\TvBox\tvbox\excel_xianlu.xlsx' #家里路径
    excel_file = r'D:\lxd\learn\py\tvbox\excel_xianlu.xlsx' #公司路径
    workbook = openpyxl.load_workbook(filename=excel_file)
    sheet1= workbook['0103']
    for row in sheet1.iter_rows(values_only=True):
        if row:
            analysis_respone=url_respon(row[1]) 
            strtxt=analysis_respone[2] #返回解析结果,如单仓/多仓/失败
            codetxt=analysis_respone[1]  #返回 url的返回状态 如200，400
            sheet1.cell(row = i,column = 3,value = strtxt)
            sheet1.cell(row = i,column = 4,value = codetxt)
            # sheet1.cell(row = i,column = 6,value = analysis_respone[0]) #返回代理链接
            print(strtxt,codetxt,row[1])
        else:
            break
        i=i+1
    workbook.save(filename = "excel_0107_2.xlsx")
    print("__"*100)
if __name__ == '__main__':
    main()
