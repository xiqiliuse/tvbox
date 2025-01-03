# 从其他py脚本中读取 单个url仓库，返回线路条数和线路

# from KuSingleURL import mainx
import KuSingleURL
import openpyxl
import requests
import json
import re,os
import warnings 
warnings.filterwarnings("ignore")

# KuSingleURL.process_json(url,"12345")

def main():
    # url="https://9877.kstore.space/FourDS/api.json"
    # url1="https://qixing.myhkw.com/DC.txt"
    # mainx(url)
    
    i=1
    # excel_file = r'D:\OneDrive\Learn\pythonLearn\TvBox\tvbox\excel_xianlu.xlsx' #家里路径
    excel_file = r'D:\lxd\learn\py\tvbox\excel_0103.xlsx' #公司路径
    workbook = openpyxl.load_workbook(filename=excel_file)
    sheet1= workbook['单仓']
    for row in sheet1.iter_rows(values_only=True):
        if not row is None :
            # analysis_respone=url_respon(row[1]) 
            returninfo=KuSingleURL.process_json(row[1],row[0])

            # strtxt=analysis_respone[2] #返回解析结果,如单仓/多仓/失败
            # codetxt=analysis_respone[1]  #返回 url的返回状态 如200，400
            # sheet1.cell(row = i,column = 4,value = strtxt)
            # sheet1.cell(row = i,column = 5,value = codetxt)

            print(row[0],row[1])
        else:
            break
        i=i+1
    # workbook.save(filename = "excel_0103sb2.xlsx")
    # print("__"*100)
if __name__ == '__main__':
    main()