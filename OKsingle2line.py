#2025-01-06更新
#从Excel读取单仓URL和名字，调用KuSingleURL.py解析，保存单仓为json文件，解析保存结果到Excel
# 从其他py脚本中读取 单个url仓库，返回线路条数和线路

# from KuSingleURL import mainx
import KuSingleURL
import openpyxl
import warnings 
warnings.filterwarnings("ignore")
def main():
    i=1
    # excel_file = r'D:\OneDrive\Learn\pythonLearn\TvBox\tvbox\excel_xianlu.xlsx' #家里路径
    excel_file = r'D:\lxd\learn\py\tvbox\excel_0103.xlsx' #公司路径
    workbook = openpyxl.load_workbook(filename=excel_file)
    sheet1= workbook['单仓']
    for row in sheet1.iter_rows(values_only=True):
        if row:
            jsonfilename = row[0] + ".json"
            returninfo=KuSingleURL.mainx(row[1],jsonfilename)
            sheet1.cell(row = i,column = 3,value = returninfo)
            print(row[0],row[1])
        else:
            break
        i=i+1
    workbook.save(filename = "excel_0107_1.xlsx")
    # print("__"*100)
if __name__ == '__main__':
    main()